#!/usr/bin/env python3
"""Load or generate lecture-hall assignment instances and solve them."""

from __future__ import annotations

import argparse
from collections import defaultdict
import datetime as dt
import json
import math
import os
import platform
import socket
import sys
import time
from dataclasses import replace
from pathlib import Path
from typing import Any

import pandas as pd
from gurobipy import GRB, GurobiError, Model, quicksum
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from ortools.sat.python import cp_model

from lecture_hall_instance_builder import FREE_WASTE_RATIO, min_students_without_waste_penalty
from lecture_hall_models import CapacityDominanceCut, Hall, Instance, Lecture
from prepare_lancs_yr23_greedy_terms import load_lancs_yr23_term_instances
from prepare_itc2019_inputs import load_itc2019_day_instances
from synthetic_instance_generator import build_synthetic_instance


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Load a lecture-hall assignment instance from a supported source, then either "
            "solve it with the selected formulation(s) or print/export the generated input."
        ),
    )
    parser.add_argument(
        "--source",
        dest="source",
        type=str,
        choices=("synthetic", "itc2019", "lancs_yr23"),
        default="synthetic",
        help="Input source. Default: synthetic.",
    )
    parser.add_argument(
        "--num-halls",
        "--num_halls",
        dest="num_halls",
        type=int,
        default=None,
        help="Number of halls for synthetic generation.",
    )
    parser.add_argument(
        "--slots-per-day",
        "--slots_per_day",
        dest="slots_per_day",
        type=int,
        default=12,
        help="Number of discrete slots per day. Default: 12.",
    )
    parser.add_argument(
        "--seed",
        dest="seed",
        type=str,
        default="0",
        help="Random seed spec for synthetic generation. Default: 0.",
    )
    parser.add_argument(
        "--density",
        dest="density",
        type=float,
        default=0.9,
        help="Target lecture-slot utilization for synthetic generation. Default: 0.9.",
    )
    parser.add_argument(
        "--itc-instance",
        dest="itc_instance",
        type=str,
        default=None,
        help="ITC 2019 instance stem, filename, or XML path.",
    )
    parser.add_argument(
        "--itc-solution",
        dest="itc_solution",
        type=Path,
        default=None,
        help="Optional ITC 2019 solution XML path.",
    )
    parser.add_argument(
        "--itc-day",
        dest="itc_day",
        type=int,
        default=None,
        help="Optional 0-based source day index for ITC 2019. When omitted, all active days are loaded.",
    )
    parser.add_argument(
        "--itc-week-index",
        dest="itc_week_index",
        type=int,
        default=None,
        help="Optional 0-based week index for ITC 2019. When omitted, the first substantial teaching week is selected.",
    )
    parser.add_argument(
        "--itc-short-break-slots",
        dest="itc_short_break_slots",
        type=int,
        default=None,
        help=(
            "Optional successor gap threshold in raw ITC slots. When omitted, the importer infers a short-break "
            "threshold automatically from student transitions."
        ),
    )
    parser.add_argument(
        "--no-capacity-fix",
        dest="itc_capacity_fix",
        action="store_false",
        help="Disable the default ITC capacity fix that reduces oversized lectures to their assigned room capacity.",
    )
    parser.add_argument(
        "--time-limit",
        "--time_limit",
        dest="time_limit",
        type=float,
        default=60.0,
        help="Per-solver time limit in seconds. Default: 60.",
    )
    parser.add_argument(
        "--cuts",
        dest="cuts",
        type=int,
        choices=(0, 1, 2, 3),
        default=1,
        help=(
            "Pair-distance cut mode: 0 = base link constraints only, "
            "1 = strong cut only, 2 = strong + symmetric strong cuts, "
            "3 = one-sided extended strong cuts. The CP model uses mode 3 as an "
            "additional propagation layer. Default: 1."
        ),
    )
    parser.add_argument(
        "--cardinality",
        dest="cardinality",
        action="store_true",
        help=(
            "Enable capacity-dominance cardinality constraints derived from maximal "
            "overlap cliques and hall-capacity thresholds. Disabled by default."
        ),
    )
    parser.add_argument(
        "--model",
        dest="model",
        type=str,
        default=None,
        help=(
            "Optional single model to solve: MIPQ, MIP, CP, or ROOT. "
            "When omitted, the script solves the three original models."
        ),
    )
    parser.add_argument(
        "--compatibility-preprocess",
        "--compatibility_preprocess",
        dest="compatibility_preprocess",
        type=str,
        choices=("none", "full", "light"),
        default="none",
        help=(
            "Optional CP-SAT preprocessing that shrinks the compatible-hall sets. "
            "Modes: none = disabled, full = maximize the target lecture hall size "
            "over all lectures, light = same maximization restricted to the target "
            "lecture and lectures that overlap it. Default: none."
        ),
    )
    parser.add_argument(
        "--instance-only",
        dest="instance_only",
        action="store_true",
        help=(
            "Generate the instance only, print the full optimization input in a "
            "human-readable terminal format, and write machine-readable JSON files. "
            "No solver is run and no Excel workbook is written."
        ),
    )
    parser.add_argument(
        "--output",
        dest="output",
        type=Path,
        default=None,
        help=(
            'Optional output workbook path. Defaults to "results.xlsx". '
            "In --instance-only mode, only the filename stem is used to name the JSON export(s)."
        ),
    )
    parser.add_argument(
        "-s",
        "--save-json",
        dest="save_json",
        action="store_true",
        help=(
            "Also write a JSON file with the full instance and all solver solutions. "
            "This is not needed in --instance-only mode because JSON export is automatic there."
        ),
    )
    parser.add_argument(
        "-q",
        "--quiet",
        action="store_true",
        help="Disable solver terminal output while running.",
    )
    args = parser.parse_args()
    validate_args(args)
    return args


def validate_args(args: argparse.Namespace) -> None:
    if args.source == "synthetic":
        if args.num_halls is None or args.num_halls <= 0:
            raise SystemExit("--num-halls must be positive when --source synthetic is used.")
        if args.slots_per_day < 4:
            raise SystemExit("--slots-per-day must be at least 4 because lectures last 2-4 slots.")
        if not 0 < args.density <= 1:
            raise SystemExit("--density must be in (0, 1].")
    elif args.source == "itc2019":
        if not args.itc_instance:
            raise SystemExit("--itc-instance is required when --source itc2019 is used.")
        if args.itc_week_index is not None and args.itc_week_index < 0:
            raise SystemExit("--itc-week-index must be nonnegative.")
        if args.itc_day is not None and args.itc_day < 0:
            raise SystemExit("--itc-day must be nonnegative.")
        if args.itc_short_break_slots is not None and args.itc_short_break_slots < 0:
            raise SystemExit("--itc-short-break-slots must be nonnegative.")
    else:
        if args.itc_week_index is not None:
            raise SystemExit("--itc-week-index is not supported when --source lancs_yr23 is used.")
        if args.itc_solution is not None:
            raise SystemExit("--itc-solution is not supported when --source lancs_yr23 is used.")
        if args.itc_day is not None and args.itc_day < 0:
            raise SystemExit("--itc-day must be nonnegative.")
        if args.itc_short_break_slots is not None and args.itc_short_break_slots < 0:
            raise SystemExit("--itc-short-break-slots must be nonnegative.")
    if args.time_limit <= 0:
        raise SystemExit("--time-limit must be positive.")
    if args.model is not None:
        args.model = args.model.upper()
        if args.model == "MIQP":
            args.model = "MIPQ"
        if args.model not in {"MIPQ", "MIP", "CP", "ROOT"}:
            raise SystemExit("--model must be one of MIPQ, MIP, CP, or ROOT.")
    if args.instance_only and args.model is not None:
        raise SystemExit("--instance-only cannot be combined with --model.")


def ensure_output_path(path: Path | None) -> Path:
    if path is None:
        return Path("results.xlsx")
    if path.suffix.lower() != ".xlsx":
        return path.with_suffix(".xlsx")
    return path


def parse_seed_range(seed_str: str) -> list[int]:
    parts = str(seed_str).split("-")
    try:
        if len(parts) == 1:
            return [int(parts[0])]
        elif len(parts) == 2:
            return list(range(int(parts[0]), int(parts[1]) + 1))
        elif len(parts) == 3:
            return list(range(int(parts[0]), int(parts[2]) + 1, int(parts[1])))
    except ValueError:
        pass
    raise ValueError(f"Invalid seed format: {seed_str}")


def build_json_path(output_path: Path, run_tag: str) -> Path:
    return output_path.with_name(f"{output_path.stem}_{run_tag}.json")


def build_instance_json_path(output_path: Path, run_tag: str) -> Path:
    return output_path.with_name(f"{output_path.stem}_instance_{run_tag}.json")


def gurobi_thread_limit() -> int:
    cpu_count = os.cpu_count() or 1
    system = platform.system()
    if system == "Darwin":
        return min(cpu_count, 8)
    if system == "Linux":
        return min(cpu_count, 12)
    return min(cpu_count, 8)


def build_overlap_neighbors(instance: Instance) -> dict[int, set[int]]:
    neighbors = {lecture.lecture_id: set() for lecture in instance.lectures}
    for active_lectures in instance.active_lectures_by_slot.values():
        if len(active_lectures) <= 1:
            continue
        for index, lecture_id_1 in enumerate(active_lectures):
            for lecture_id_2 in active_lectures[index + 1 :]:
                neighbors[lecture_id_1].add(lecture_id_2)
                neighbors[lecture_id_2].add(lecture_id_1)
    return neighbors


def build_maximal_active_cliques(instance: Instance) -> list[tuple[int, tuple[int, ...]]]:
    lecture_map = {lecture.lecture_id: lecture for lecture in instance.lectures}
    active_sets_by_day: dict[int, set[frozenset[int]]] = defaultdict(set)

    for active_lectures in instance.active_lectures_by_slot.values():
        if len(active_lectures) <= 1:
            continue
        active_set = frozenset(active_lectures)
        sample_lecture_id = next(iter(active_set))
        day = lecture_map[sample_lecture_id].day
        active_sets_by_day[day].add(active_set)

    maximal_cliques: list[tuple[int, tuple[int, ...]]] = []
    for day in sorted(active_sets_by_day):
        day_sets = sorted(
            active_sets_by_day[day],
            key=lambda active_set: (
                -len(active_set),
                min(lecture_map[lecture_id].start_slot for lecture_id in active_set),
                tuple(sorted(active_set)),
            ),
        )
        maximal_sets: list[frozenset[int]] = []
        for active_set in day_sets:
            if any(active_set < other_set for other_set in maximal_sets):
                continue
            maximal_sets.append(active_set)

        maximal_sets.sort(
            key=lambda active_set: (
                min(lecture_map[lecture_id].start_slot for lecture_id in active_set),
                tuple(sorted(active_set)),
            )
        )
        maximal_cliques.extend(
            (day, tuple(sorted(active_set)))
            for active_set in maximal_sets
        )

    return maximal_cliques


def build_capacity_dominance_cuts(instance: Instance) -> list[CapacityDominanceCut]:
    lecture_map = {lecture.lecture_id: lecture for lecture in instance.lectures}
    threshold_halls: list[tuple[int, tuple[int, ...], set[int]]] = []
    for threshold in sorted({hall.capacity for hall in instance.halls}):
        large_hall_ids = tuple(
            sorted(hall.hall_id for hall in instance.halls if hall.capacity > threshold)
        )
        if not large_hall_ids:
            continue
        threshold_halls.append((threshold, large_hall_ids, set(large_hall_ids)))

    cuts: list[CapacityDominanceCut] = []
    for clique_index, (day, clique_lecture_ids) in enumerate(build_maximal_active_cliques(instance)):
        for threshold, large_hall_ids, large_hall_id_set in threshold_halls:
            num_large_lectures = sum(
                1
                for lecture_id in clique_lecture_ids
                if lecture_map[lecture_id].students > threshold
            )
            rhs = len(large_hall_ids) - num_large_lectures
            eligible_lecture_ids = tuple(
                lecture_id
                for lecture_id in clique_lecture_ids
                if lecture_map[lecture_id].students <= threshold
                and any(
                    hall_id in large_hall_id_set
                    for hall_id in instance.compatibility[lecture_id]
                )
            )
            if not eligible_lecture_ids and rhs >= 0:
                continue
            if rhs >= len(eligible_lecture_ids):
                continue
            cuts.append(
                CapacityDominanceCut(
                    clique_index=clique_index,
                    day=day,
                    threshold=threshold,
                    eligible_lecture_ids=eligible_lecture_ids,
                    large_hall_ids=large_hall_ids,
                    rhs=rhs,
                )
            )

    return cuts


def add_cp_extended_strong_distance_propagation(
    model: cp_model.CpModel,
    instance: Instance,
    hall_assignment: dict[int, cp_model.IntVar],
    z_vars: dict[tuple[int, int], cp_model.IntVar],
) -> None:
    subset_indicator_cache: dict[tuple[int, tuple[int, ...]], cp_model.IntVar] = {}

    def subset_indicator(lecture_id: int, hall_subset: tuple[int, ...]) -> cp_model.IntVar:
        key = (lecture_id, hall_subset)
        if key in subset_indicator_cache:
            return subset_indicator_cache[key]

        indicator = model.NewBoolVar(f"zcut_{lecture_id}_{len(subset_indicator_cache)}")
        hall_subset_set = set(hall_subset)
        model.AddAllowedAssignments(
            [hall_assignment[lecture_id], indicator],
            [
                (hall_id, 1 if hall_id in hall_subset_set else 0)
                for hall_id in instance.compatibility[lecture_id]
            ],
        )
        subset_indicator_cache[key] = indicator
        return indicator

    for (lecture_id_1, lecture_id_2), z_var in z_vars.items():
        halls_1 = instance.compatibility[lecture_id_1]
        halls_2 = instance.compatibility[lecture_id_2]
        seen_patterns: set[tuple[tuple[int, ...], tuple[int, ...], int]] = set()

        for hall_id_1 in halls_1:
            for hall_id_2 in halls_2:
                threshold_distance = instance.distances[hall_id_1][hall_id_2]
                far_halls_2 = tuple(
                    hall_id
                    for hall_id in halls_2
                    if instance.distances[hall_id_1][hall_id] >= threshold_distance
                )
                far_halls_1 = tuple(
                    hall_id
                    for hall_id in halls_1
                    if all(
                        instance.distances[hall_id][far_hall_2] >= threshold_distance
                        for far_hall_2 in far_halls_2
                    )
                )
                pattern = (far_halls_1, far_halls_2, threshold_distance)
                if pattern in seen_patterns:
                    continue
                seen_patterns.add(pattern)

                left_indicator = subset_indicator(lecture_id_1, far_halls_1)
                right_indicator = subset_indicator(lecture_id_2, far_halls_2)
                model.Add(z_var >= threshold_distance).OnlyEnforceIf(
                    [left_indicator, right_indicator]
                )


def cp_sat_capacity_upper_bound(
    instance: Instance,
    compatibility: dict[int, list[int]],
    target_lecture_id: int,
    lecture_subset: list[int],
    num_search_workers: int,
) -> tuple[int | None, str]:
    model = cp_model.CpModel()
    lecture_subset_set = set(lecture_subset)
    hall_assignment: dict[int, cp_model.IntVar] = {}

    for lecture_id in lecture_subset:
        compatible_halls = compatibility[lecture_id]
        if not compatible_halls:
            return None, "INFEASIBLE"
        hall_assignment[lecture_id] = model.NewIntVarFromDomain(
            cp_model.Domain.FromValues(compatible_halls),
            f"pre_a_{lecture_id}",
        )

    for active_lectures in instance.active_lectures_by_slot.values():
        scoped_active_lectures = [lecture_id for lecture_id in active_lectures if lecture_id in lecture_subset_set]
        if len(scoped_active_lectures) > 1:
            model.AddAllDifferent([hall_assignment[lecture_id] for lecture_id in scoped_active_lectures])

    hall_capacity_by_id = {hall.hall_id: hall.capacity for hall in instance.halls}
    compatible_capacities = [
        hall_capacity_by_id[hall_id] for hall_id in compatibility[target_lecture_id]
    ]
    target_capacity = model.NewIntVar(
        min(compatible_capacities),
        max(compatible_capacities),
        f"pre_cap_{target_lecture_id}",
    )
    model.AddAllowedAssignments(
        [hall_assignment[target_lecture_id], target_capacity],
        [
            (hall_id, hall_capacity_by_id[hall_id])
            for hall_id in compatibility[target_lecture_id]
        ],
    )
    model.Maximize(target_capacity)

    solver = cp_model.CpSolver()
    solver.parameters.num_search_workers = num_search_workers
    solver.parameters.random_seed = instance.seed
    solver.parameters.log_search_progress = False
    solver.parameters.catch_sigint_signal = False
    status = solver.Solve(model)
    status_name = status_name_from_cp_sat(status)
    if status == cp_model.INFEASIBLE:
        return None, status_name

    upper_capacity = max(compatible_capacities)
    best_bound = safe_float(solver.BestObjectiveBound())
    if best_bound is not None:
        upper_capacity = min(upper_capacity, math.floor(best_bound + 1e-9))

    incumbent_capacity = safe_float(solver.ObjectiveValue())
    if incumbent_capacity is not None:
        upper_capacity = max(upper_capacity, int(round(incumbent_capacity)))

    upper_capacity = max(upper_capacity, min(compatible_capacities))
    return upper_capacity, status_name


def apply_compatibility_preprocessing(
    instance: Instance,
    mode: str,
) -> tuple[Instance, bool]:
    compatibility_entries_before = sum(len(hall_ids) for hall_ids in instance.compatibility.values())
    thread_limit = gurobi_thread_limit()
    if mode == "none":
        return (
            replace(
                instance,
                compatibility_preprocess_mode="none",
                compatibility_entries_before=compatibility_entries_before,
                compatibility_entries_after=compatibility_entries_before,
                compatibility_preprocess_subproblems=0,
                compatibility_preprocess_wall_seconds=0.0,
                compatibility_preprocess_tightened_lectures=0,
                compatibility_preprocess_optimal_subproblems=0,
                compatibility_preprocess_nonoptimal_subproblems=0,
            ),
            False,
        )

    overlap_neighbors = build_overlap_neighbors(instance)
    compatibility = {
        lecture_id: list(hall_ids)
        for lecture_id, hall_ids in instance.compatibility.items()
    }
    hall_capacity_by_id = {hall.hall_id: hall.capacity for hall in instance.halls}
    lecture_ids_all = [lecture.lecture_id for lecture in instance.lectures]
    wall_start = time.perf_counter()
    tightened_lecture_ids: set[int] = set()
    total_subproblems = 0
    optimal_subproblems = 0
    nonoptimal_subproblems = 0

    infeasible = False
    while True:
        reduced_compatibility: dict[int, list[int]] = {}
        round_changed = False
        infeasible = False

        for lecture in instance.lectures:
            lecture_id = lecture.lecture_id
            if mode == "full":
                lecture_subset = lecture_ids_all
            else:
                lecture_subset = sorted({lecture_id, *overlap_neighbors[lecture_id]})

            upper_capacity, status_name = cp_sat_capacity_upper_bound(
                instance=instance,
                compatibility=compatibility,
                target_lecture_id=lecture_id,
                lecture_subset=lecture_subset,
                num_search_workers=thread_limit,
            )
            total_subproblems += 1
            if status_name == "OPTIMAL":
                optimal_subproblems += 1
            else:
                nonoptimal_subproblems += 1
            if upper_capacity is None:
                reduced_compatibility[lecture_id] = []
                infeasible = True
                continue

            max_capacity_before = max(hall_capacity_by_id[hall_id] for hall_id in compatibility[lecture_id])
            reduced_compatibility[lecture_id] = [
                hall_id
                for hall_id in compatibility[lecture_id]
                if hall_capacity_by_id[hall_id] <= upper_capacity
            ]
            if not reduced_compatibility[lecture_id]:
                infeasible = True
            if reduced_compatibility[lecture_id] != compatibility[lecture_id]:
                round_changed = True
                if upper_capacity < max_capacity_before:
                    tightened_lecture_ids.add(lecture_id)

        compatibility = {
            lecture_id: list(hall_ids)
            for lecture_id, hall_ids in reduced_compatibility.items()
        }
        if infeasible or not round_changed:
            break

    reduced_assignment_penalties = {
        lecture_id: {
            hall_id: instance.assignment_penalties[lecture_id][hall_id]
            for hall_id in reduced_compatibility[lecture_id]
        }
        for lecture_id in reduced_compatibility
    }
    compatibility_entries_after = sum(len(hall_ids) for hall_ids in reduced_compatibility.values())
    return (
        replace(
            instance,
            compatibility=reduced_compatibility,
            assignment_penalties=reduced_assignment_penalties,
            compatibility_preprocess_mode=mode,
            compatibility_entries_before=compatibility_entries_before,
            compatibility_entries_after=compatibility_entries_after,
            compatibility_preprocess_subproblems=total_subproblems,
            compatibility_preprocess_wall_seconds=time.perf_counter() - wall_start,
            compatibility_preprocess_tightened_lectures=len(tightened_lecture_ids),
            compatibility_preprocess_optimal_subproblems=optimal_subproblems,
            compatibility_preprocess_nonoptimal_subproblems=nonoptimal_subproblems,
        ),
        infeasible,
    )


def count_decomposition_connected_components(instance: Instance) -> int:
    adjacency: dict[int, set[int]] = {
        lecture.lecture_id: set() for lecture in instance.lectures
    }

    # Successor arcs induce objective coupling.
    for lecture_id_1, lecture_id_2 in instance.common_students:
        adjacency[lecture_id_1].add(lecture_id_2)
        adjacency[lecture_id_2].add(lecture_id_1)

    # Overlap edges induce room-competition coupling.
    for active_lectures in instance.active_lectures_by_slot.values():
        if len(active_lectures) <= 1:
            continue
        for index, lecture_id_1 in enumerate(active_lectures):
            for lecture_id_2 in active_lectures[index + 1 :]:
                adjacency[lecture_id_1].add(lecture_id_2)
                adjacency[lecture_id_2].add(lecture_id_1)

    visited: set[int] = set()
    component_count = 0
    for lecture in instance.lectures:
        lecture_id = lecture.lecture_id
        if lecture_id in visited:
            continue

        component_count += 1
        stack = [lecture_id]
        visited.add(lecture_id)
        while stack:
            current = stack.pop()
            for neighbor in adjacency[current]:
                if neighbor in visited:
                    continue
                visited.add(neighbor)
                stack.append(neighbor)

    return component_count


def safe_float(value: Any) -> float | None:
    try:
        value = float(value)
    except (TypeError, ValueError):
        return None
    if math.isfinite(value):
        return value
    return None


def status_name_from_gurobi(status_code: int) -> str:
    mapping = {
        GRB.LOADED: "LOADED",
        GRB.OPTIMAL: "OPTIMAL",
        GRB.INFEASIBLE: "INFEASIBLE",
        GRB.INF_OR_UNBD: "INF_OR_UNBD",
        GRB.UNBOUNDED: "UNBOUNDED",
        GRB.CUTOFF: "CUTOFF",
        GRB.ITERATION_LIMIT: "ITERATION_LIMIT",
        GRB.NODE_LIMIT: "NODE_LIMIT",
        GRB.TIME_LIMIT: "TIME_LIMIT",
        GRB.SOLUTION_LIMIT: "SOLUTION_LIMIT",
        GRB.INTERRUPTED: "INTERRUPTED",
        GRB.NUMERIC: "NUMERIC",
        GRB.SUBOPTIMAL: "SUBOPTIMAL",
        GRB.USER_OBJ_LIMIT: "USER_OBJ_LIMIT",
    }
    return mapping.get(status_code, f"STATUS_{status_code}")


def status_name_from_cp_sat(status_code: int) -> str:
    mapping = {
        cp_model.OPTIMAL: "OPTIMAL",
        cp_model.FEASIBLE: "FEASIBLE",
        cp_model.INFEASIBLE: "INFEASIBLE",
        cp_model.MODEL_INVALID: "MODEL_INVALID",
        cp_model.UNKNOWN: "UNKNOWN",
    }
    return mapping.get(status_code, f"STATUS_{status_code}")


def assignment_details_from_map(
    instance: Instance,
    assignment_by_lecture: dict[int, int] | None,
) -> dict[str, Any] | None:
    if assignment_by_lecture is None:
        return None

    hall_map = {hall.hall_id: hall for hall in instance.halls}
    lecture_map = {lecture.lecture_id: lecture for lecture in instance.lectures}
    assignments = []
    walking_objective_terms = []
    assignment_penalty_terms = []
    recomputed_walking_objective = 0
    recomputed_assignment_penalty = 0

    for lecture in instance.lectures:
        hall_id = assignment_by_lecture[lecture.lecture_id]
        hall = hall_map[hall_id]
        unused_seats = hall.capacity - lecture.students
        min_students_free = min_students_without_waste_penalty(hall.capacity)
        allowed_empty_seats = hall.capacity - min_students_free
        excess_empty_seats = max(0, unused_seats - allowed_empty_seats)
        assignment_penalty = instance.assignment_penalties[lecture.lecture_id][hall_id]
        recomputed_assignment_penalty += assignment_penalty
        assignments.append(
            {
                "lecture_id": lecture.lecture_id,
                "lecture_name": lecture.name,
                "subject": lecture.subject,
                "study_year": lecture.study_year,
                "course_type": "compulsory" if lecture.is_compulsory else "elective",
                "day": lecture.day,
                "start_slot": lecture.start_slot,
                "end_slot": lecture.end_slot,
                "students": lecture.students,
                "assigned_hall_id": hall_id,
                "assigned_hall_name": hall.name,
                "assigned_hall_capacity": hall.capacity,
                "unused_seats": unused_seats,
                "allowed_empty_seats_without_penalty": allowed_empty_seats,
                "excess_empty_seats_penalized": excess_empty_seats,
                "assignment_penalty": assignment_penalty,
            }
        )
        assignment_penalty_terms.append(
            {
                "term_type": "assignment_penalty",
                "lecture_id": lecture.lecture_id,
                "lecture_name": lecture.name,
                "hall_id": hall_id,
                "hall_name": hall.name,
                "hall_capacity": hall.capacity,
                "students": lecture.students,
                "unused_seats": unused_seats,
                "allowed_empty_seats_without_penalty": allowed_empty_seats,
                "excess_empty_seats_penalized": excess_empty_seats,
                "contribution": assignment_penalty,
            }
        )

    for (lecture_id_1, lecture_id_2), common_count in sorted(instance.common_students.items()):
        hall_id_1 = assignment_by_lecture[lecture_id_1]
        hall_id_2 = assignment_by_lecture[lecture_id_2]
        distance = instance.distances[hall_id_1][hall_id_2]
        contribution = common_count * distance
        recomputed_walking_objective += contribution
        walking_objective_terms.append(
            {
                "term_type": "walking",
                "from_lecture_id": lecture_id_1,
                "from_lecture_name": lecture_map[lecture_id_1].name,
                "to_lecture_id": lecture_id_2,
                "to_lecture_name": lecture_map[lecture_id_2].name,
                "common_students": common_count,
                "from_hall_id": hall_id_1,
                "to_hall_id": hall_id_2,
                "distance": distance,
                "contribution": contribution,
            }
        )

    return {
        "assignment_by_lecture": assignment_by_lecture,
        "assignments": assignments,
        "objective_terms": walking_objective_terms + assignment_penalty_terms,
        "walking_objective_terms": walking_objective_terms,
        "assignment_penalty_terms": assignment_penalty_terms,
        "recomputed_walking_objective": recomputed_walking_objective,
        "recomputed_assignment_penalty": recomputed_assignment_penalty,
        "recomputed_objective": recomputed_walking_objective + recomputed_assignment_penalty,
    }


def build_gurobi_linearized_model(
    instance: Instance,
    cuts: int,
    time_limit: float | None,
    verbose: bool,
    cardinality: bool = False,
    threads: int | None = None,
) -> tuple[Model, dict[tuple[int, int], Any], dict[tuple[int, int], Any], int]:
    thread_limit = gurobi_thread_limit() if threads is None else threads
    model = Model("lecture_hall_linearized")
    model.Params.OutputFlag = 1 if verbose else 0
    if time_limit is not None:
        model.Params.TimeLimit = time_limit
    model.Params.Threads = thread_limit
    if cuts == 0:  # Otherwise the problem frequently ends with NUMERIC error prematurely
        model.Params.NumericFocus = 2
        model.Params.ScaleFlag = 2
        model.Params.ObjScale = -0.5

    x: dict[tuple[int, int], Any] = {}
    # The paper denotes this pair auxiliary uniformly by z_{l1,l2}. For cuts > 0
    # the implementation keeps the same z-based name even though the common-student
    # weight is folded into the strengthened linear inequalities for scaling.
    z_vars: dict[tuple[int, int], Any] = {}

    for lecture in instance.lectures:
        for hall_id in instance.compatibility[lecture.lecture_id]:
            x[(lecture.lecture_id, hall_id)] = model.addVar(
                vtype=GRB.BINARY,
                name=f"x_{lecture.lecture_id}_{hall_id}",
            )

    for lecture_id_1, lecture_id_2 in instance.common_students:
        z_vars[(lecture_id_1, lecture_id_2)] = model.addVar(
            lb=0.0,
            vtype=GRB.CONTINUOUS,
            name=f"z_{lecture_id_1}_{lecture_id_2}",
        )

    model.update()

    for lecture in instance.lectures:
        model.addConstr(
            quicksum(
                x[(lecture.lecture_id, hall_id)]
                for hall_id in instance.compatibility[lecture.lecture_id]
            )
            == 1,
            name=f"assign_{lecture.lecture_id}",
        )

    for slot, active_lectures in instance.active_lectures_by_slot.items():
        if len(active_lectures) <= 1:
            continue
        for hall in instance.halls:
            vars_for_slot = [
                x[(lecture_id, hall.hall_id)]
                for lecture_id in active_lectures
                if (lecture_id, hall.hall_id) in x
            ]
            if len(vars_for_slot) > 1:
                model.addConstr(
                    quicksum(vars_for_slot) <= 1,
                    name=f"overlap_h{hall.hall_id}_t{slot}",
                )

    if cardinality:
        for cut in build_capacity_dominance_cuts(instance):
            expr = quicksum(
                x[(lecture_id, hall_id)]
                for lecture_id in cut.eligible_lecture_ids
                for hall_id in cut.large_hall_ids
                if (lecture_id, hall_id) in x
            )
            model.addConstr(
                expr <= cut.rhs,
                name=f"card_d{cut.day}_c{cut.clique_index}_k{cut.threshold}",
            )

    if cuts == 0:
        for lecture_id_1, lecture_id_2 in instance.common_students:
            for hall_id_1 in instance.compatibility[lecture_id_1]:
                for hall_id_2 in instance.compatibility[lecture_id_2]:
                    model.addConstr(
                        z_vars[(lecture_id_1, lecture_id_2)]
                        >= instance.distances[hall_id_1][hall_id_2]
                        * (x[(lecture_id_1, hall_id_1)] + x[(lecture_id_2, hall_id_2)] - 1),
                        name=f"link_{lecture_id_1}_{lecture_id_2}_{hall_id_1}_{hall_id_2}",
                    )

    if cuts in (1, 2):
        for (lecture_id_1, lecture_id_2), common_count in instance.common_students.items():
            halls_1 = instance.compatibility[lecture_id_1]
            halls_2 = instance.compatibility[lecture_id_2]

            for hall_id_1 in halls_1:
                for hall_id_2 in halls_2:
                    threshold_distance = instance.distances[hall_id_1][hall_id_2]
                    farther_halls = [
                        hall_id
                        for hall_id in halls_2
                        if instance.distances[hall_id_1][hall_id] >= threshold_distance
                    ]
                    model.addConstr(
                        z_vars[(lecture_id_1, lecture_id_2)]
                        >= common_count
                        * threshold_distance
                        * (
                            x[(lecture_id_1, hall_id_1)]
                            - 1
                            + quicksum(x[(lecture_id_2, hall_id)] for hall_id in farther_halls)
                        ),
                        name=f"strong_{lecture_id_1}_{lecture_id_2}_{hall_id_1}_{hall_id_2}",
                    )

    if cuts == 2:
        for (lecture_id_1, lecture_id_2), common_count in instance.common_students.items():
            halls_1 = instance.compatibility[lecture_id_1]
            halls_2 = instance.compatibility[lecture_id_2]

            for hall_id_1 in halls_1:
                for hall_id_2 in halls_2:
                    threshold_distance = instance.distances[hall_id_1][hall_id_2]
                    farther_halls = [
                        hall_id
                        for hall_id in halls_1
                        if instance.distances[hall_id][hall_id_2] >= threshold_distance
                    ]
                    model.addConstr(
                        z_vars[(lecture_id_1, lecture_id_2)]
                        >= common_count
                        * threshold_distance
                        * (
                            quicksum(x[(lecture_id_1, hall_id)] for hall_id in farther_halls)
                            - 1
                            + x[(lecture_id_2, hall_id_2)]
                        ),
                        name=f"strongsym_{lecture_id_1}_{lecture_id_2}_{hall_id_1}_{hall_id_2}",
                    )

    if cuts == 3:
        for (lecture_id_1, lecture_id_2), common_count in instance.common_students.items():
            halls_1 = instance.compatibility[lecture_id_1]
            halls_2 = instance.compatibility[lecture_id_2]

            for hall_id_1 in halls_1:
                for hall_id_2 in halls_2:
                    threshold_distance = instance.distances[hall_id_1][hall_id_2]
                    far_halls_2 = [
                        hall_id
                        for hall_id in halls_2
                        if instance.distances[hall_id_1][hall_id] >= threshold_distance
                    ]
                    far_halls_1 = [
                        hall_id
                        for hall_id in halls_1
                        if all(
                            instance.distances[hall_id][far_hall_2] >= threshold_distance
                            for far_hall_2 in far_halls_2
                        )
                    ]
                    model.addConstr(
                        z_vars[(lecture_id_1, lecture_id_2)]
                        >= common_count
                        * threshold_distance
                        * (
                            quicksum(x[(lecture_id_1, hall_id)] for hall_id in far_halls_1)
                            - 1
                            + quicksum(x[(lecture_id_2, hall_id)] for hall_id in far_halls_2)
                        ),
                        name=f"strongext_{lecture_id_1}_{lecture_id_2}_{hall_id_1}_{hall_id_2}",
                    )

    if cuts == 0:
        # Keep the pair weights in the objective to improve matrix scaling.
        objective_terms = [
            common_count * z_vars[(lecture_id_1, lecture_id_2)]
            for (lecture_id_1, lecture_id_2), common_count in instance.common_students.items()
        ]
    else:
        objective_terms = [
            z_vars[(lecture_id_1, lecture_id_2)]
            for (lecture_id_1, lecture_id_2) in instance.common_students
        ]
    objective_terms.extend(
        instance.assignment_penalties[lecture.lecture_id][hall_id] * x[(lecture.lecture_id, hall_id)]
        for lecture in instance.lectures
        for hall_id in instance.compatibility[lecture.lecture_id]
    )
    model.setObjective(quicksum(objective_terms), GRB.MINIMIZE)
    return model, x, z_vars, thread_limit


def solve_gurobi_quadratic(
    instance: Instance,
    time_limit: float,
    verbose: bool = True,
    cardinality: bool = False,
) -> dict[str, Any]:
    wall_start = time.perf_counter()
    thread_limit = gurobi_thread_limit()
    try:
        model = Model("lecture_hall_quadratic")
        model.Params.OutputFlag = 1 if verbose else 0
        model.Params.TimeLimit = time_limit
        model.Params.Threads = thread_limit
        model.Params.NonConvex = 2

        x: dict[tuple[int, int], Any] = {}
        for lecture in instance.lectures:
            for hall_id in instance.compatibility[lecture.lecture_id]:
                x[(lecture.lecture_id, hall_id)] = model.addVar(
                    vtype=GRB.BINARY,
                    name=f"x_{lecture.lecture_id}_{hall_id}",
                )

        model.update()

        for lecture in instance.lectures:
            model.addConstr(
                quicksum(
                    x[(lecture.lecture_id, hall_id)]
                    for hall_id in instance.compatibility[lecture.lecture_id]
                )
                == 1,
                name=f"assign_{lecture.lecture_id}",
            )

        for slot, active_lectures in instance.active_lectures_by_slot.items():
            if len(active_lectures) <= 1:
                continue
            for hall in instance.halls:
                vars_for_slot = [
                    x[(lecture_id, hall.hall_id)]
                    for lecture_id in active_lectures
                    if (lecture_id, hall.hall_id) in x
                ]
                if len(vars_for_slot) > 1:
                    model.addConstr(
                        quicksum(vars_for_slot) <= 1,
                        name=f"overlap_h{hall.hall_id}_t{slot}",
                    )

        if cardinality:
            for cut in build_capacity_dominance_cuts(instance):
                expr = quicksum(
                    x[(lecture_id, hall_id)]
                    for lecture_id in cut.eligible_lecture_ids
                    for hall_id in cut.large_hall_ids
                    if (lecture_id, hall_id) in x
                )
                model.addConstr(
                    expr <= cut.rhs,
                    name=f"card_d{cut.day}_c{cut.clique_index}_k{cut.threshold}",
                )

        objective_terms = []
        for (lecture_id_1, lecture_id_2), common_count in instance.common_students.items():
            for hall_id_1 in instance.compatibility[lecture_id_1]:
                for hall_id_2 in instance.compatibility[lecture_id_2]:
                    distance = instance.distances[hall_id_1][hall_id_2]
                    objective_terms.append(
                        common_count
                        * distance
                        * x[(lecture_id_1, hall_id_1)]
                        * x[(lecture_id_2, hall_id_2)]
                    )
        objective_terms.extend(
            instance.assignment_penalties[lecture.lecture_id][hall_id] * x[(lecture.lecture_id, hall_id)]
            for lecture in instance.lectures
            for hall_id in instance.compatibility[lecture.lecture_id]
        )
        model.setObjective(quicksum(objective_terms), GRB.MINIMIZE)
        model.optimize()

        if model.Status == GRB.INTERRUPTED:
            raise KeyboardInterrupt("Solver interrupted by user")

        wall_seconds = time.perf_counter() - wall_start
        assignment_by_lecture = None
        if model.SolCount > 0:
            assignment_by_lecture = {}
            for lecture in instance.lectures:
                assigned_hall = next(
                    hall_id
                    for hall_id in instance.compatibility[lecture.lecture_id]
                    if x[(lecture.lecture_id, hall_id)].X > 0.5
                )
                assignment_by_lecture[lecture.lecture_id] = assigned_hall
        return {
            "solver_family": "GUROBI",
            "formulation": "quadratic_miqp",
            "status": status_name_from_gurobi(model.Status),
            "objective_value": safe_float(model.ObjVal) if model.SolCount > 0 else None,
            "lower_bound": safe_float(model.ObjBound),
            "wall_clock_seconds": wall_seconds,
            "solver_runtime_seconds": safe_float(model.Runtime),
            "mip_gap": safe_float(model.MIPGap) if model.SolCount > 0 else None,
            "threads": thread_limit,
            "error": None,
            "solution": assignment_details_from_map(instance, assignment_by_lecture),
        }
    except GurobiError as error:
        return {
            "solver_family": "GUROBI",
            "formulation": "quadratic_miqp",
            "status": "ERROR",
            "objective_value": None,
            "lower_bound": None,
            "wall_clock_seconds": time.perf_counter() - wall_start,
            "solver_runtime_seconds": None,
            "mip_gap": None,
            "threads": thread_limit,
            "error": str(error),
            "solution": None,
        }


def solve_gurobi_linearized(
    instance: Instance,
    time_limit: float,
    cuts: int = 1,
    verbose: bool = True,
    cardinality: bool = False,
) -> dict[str, Any]:
    wall_start = time.perf_counter()
    try:
        model, x, _, thread_limit = build_gurobi_linearized_model(
            instance=instance,
            cuts=cuts,
            time_limit=time_limit,
            verbose=verbose,
            cardinality=cardinality,
        )
        model.optimize()

        if model.Status == GRB.INTERRUPTED:
            raise KeyboardInterrupt("Solver interrupted by user")

        wall_seconds = time.perf_counter() - wall_start
        assignment_by_lecture = None
        if model.SolCount > 0:
            assignment_by_lecture = {}
            for lecture in instance.lectures:
                assigned_hall = next(
                    hall_id
                    for hall_id in instance.compatibility[lecture.lecture_id]
                    if x[(lecture.lecture_id, hall_id)].X > 0.5
                )
                assignment_by_lecture[lecture.lecture_id] = assigned_hall
        return {
            "solver_family": "GUROBI",
            "formulation": "linearized_milp",
            "status": status_name_from_gurobi(model.Status),
            "objective_value": safe_float(model.ObjVal) if model.SolCount > 0 else None,
            "lower_bound": safe_float(model.ObjBound),
            "wall_clock_seconds": wall_seconds,
            "solver_runtime_seconds": safe_float(model.Runtime),
            "mip_gap": safe_float(model.MIPGap) if model.SolCount > 0 else None,
            "threads": thread_limit,
            "cuts_mode": cuts,
            "error": None,
            "solution": assignment_details_from_map(instance, assignment_by_lecture),
        }
    except GurobiError as error:
        return {
            "solver_family": "GUROBI",
            "formulation": "linearized_milp",
            "status": "ERROR",
            "objective_value": None,
            "lower_bound": None,
            "wall_clock_seconds": time.perf_counter() - wall_start,
            "solver_runtime_seconds": None,
            "mip_gap": None,
            "threads": thread_limit,
            "cuts_mode": cuts,
            "error": str(error),
            "solution": None,
        }


def solve_gurobi_linearized_root(
    instance: Instance,
    time_limit: float,
    cuts: int = 1,
    verbose: bool = True,
    cardinality: bool = False,
) -> dict[str, Any]:
    wall_start = time.perf_counter()
    try:
        model, _, _, thread_limit = build_gurobi_linearized_model(
            instance=instance,
            cuts=cuts,
            time_limit=time_limit,
            verbose=verbose,
            cardinality=cardinality,
        )
        model._root_bound = None
        model._terminated_after_root = False

        def callback(cb_model: Model, where: int) -> None:
            if where != GRB.Callback.MIP:
                return

            node_count = cb_model.cbGet(GRB.Callback.MIP_NODCNT)
            obj_bound = cb_model.cbGet(GRB.Callback.MIP_OBJBND)
            if node_count == 0:
                if cb_model._root_bound is None or obj_bound > cb_model._root_bound:
                    cb_model._root_bound = obj_bound
            elif not cb_model._terminated_after_root:
                cb_model._terminated_after_root = True
                cb_model.terminate()

        model.optimize(callback)

        wall_seconds = time.perf_counter() - wall_start
        root_bound = safe_float(model._root_bound)
        if root_bound is None:
            root_bound = safe_float(model.ObjBound)

        if model._terminated_after_root:
            status = "ROOT_LIMIT"
        else:
            status = status_name_from_gurobi(model.Status)

        return {
            "solver_family": "GUROBI",
            "formulation": "linearized_root",
            "status": status,
            "objective_value": None,
            "lower_bound": root_bound,
            "wall_clock_seconds": wall_seconds,
            "solver_runtime_seconds": safe_float(model.Runtime),
            "mip_gap": None,
            "threads": thread_limit,
            "cuts_mode": cuts,
            "error": None,
            "solution": None,
        }
    except GurobiError as error:
        return {
            "solver_family": "GUROBI",
            "formulation": "linearized_root",
            "status": "ERROR",
            "objective_value": None,
            "lower_bound": None,
            "wall_clock_seconds": time.perf_counter() - wall_start,
            "solver_runtime_seconds": None,
            "mip_gap": None,
            "threads": gurobi_thread_limit(),
            "cuts_mode": cuts,
            "error": str(error),
            "solution": None,
        }


def solve_cp_sat(
    instance: Instance,
    time_limit: float,
    verbose: bool = True,
    cuts: int = 1,
    cardinality: bool = False,
) -> dict[str, Any]:
    wall_start = time.perf_counter()
    model = cp_model.CpModel()
    thread_limit = gurobi_thread_limit()

    hall_assignment: dict[int, cp_model.IntVar] = {}
    for lecture in instance.lectures:
        domain = cp_model.Domain.FromValues(instance.compatibility[lecture.lecture_id])
        hall_assignment[lecture.lecture_id] = model.NewIntVarFromDomain(
            domain,
            f"a_{lecture.lecture_id}",
        )

    for active_lectures in instance.active_lectures_by_slot.values():
        if len(active_lectures) > 1:
            model.AddAllDifferent([hall_assignment[lecture_id] for lecture_id in active_lectures])

    if cardinality:
        cardinality_cuts = build_capacity_dominance_cuts(instance)
        large_halls_by_threshold = {
            cut.threshold: set(cut.large_hall_ids)
            for cut in cardinality_cuts
        }
        assigned_to_large_hall: dict[tuple[int, int], cp_model.IntVar] = {}

        def large_hall_indicator(lecture_id: int, threshold: int) -> cp_model.IntVar:
            key = (lecture_id, threshold)
            if key in assigned_to_large_hall:
                return assigned_to_large_hall[key]
            indicator = model.NewBoolVar(f"card_{lecture_id}_{threshold}")
            model.AddAllowedAssignments(
                [hall_assignment[lecture_id], indicator],
                [
                    (hall_id, 1 if hall_id in large_halls_by_threshold[threshold] else 0)
                    for hall_id in instance.compatibility[lecture_id]
                ],
            )
            assigned_to_large_hall[key] = indicator
            return indicator

        for cut in cardinality_cuts:
            if not cut.eligible_lecture_ids:
                if cut.rhs < 0:
                    model.AddBoolOr([])
                continue
            model.Add(
                sum(
                    large_hall_indicator(lecture_id, cut.threshold)
                    for lecture_id in cut.eligible_lecture_ids
                )
                <= cut.rhs
            )

    objective_terms: list[Any] = []
    z_vars: dict[tuple[int, int], cp_model.IntVar] = {}
    for lecture in instance.lectures:
        compatible_penalties = instance.assignment_penalties[lecture.lecture_id]
        feasible_penalties = sorted(set(compatible_penalties.values()))
        penalty_var = model.NewIntVar(
            feasible_penalties[0],
            feasible_penalties[-1],
            f"p_{lecture.lecture_id}",
        )
        model.AddAllowedAssignments(
            [hall_assignment[lecture.lecture_id], penalty_var],
            [
                (hall_id, compatible_penalties[hall_id])
                for hall_id in instance.compatibility[lecture.lecture_id]
            ],
        )
        objective_terms.append(penalty_var)
    for (lecture_id_1, lecture_id_2), common_count in instance.common_students.items():
        feasible_distances = sorted(
            {
                instance.distances[hall_id_1][hall_id_2]
                for hall_id_1 in instance.compatibility[lecture_id_1]
                for hall_id_2 in instance.compatibility[lecture_id_2]
            }
        )
        z_var = model.NewIntVar(
            feasible_distances[0],
            feasible_distances[-1],
            f"z_{lecture_id_1}_{lecture_id_2}",
        )
        z_vars[(lecture_id_1, lecture_id_2)] = z_var
        tuples = [
            (hall_id_1, hall_id_2, instance.distances[hall_id_1][hall_id_2])
            for hall_id_1 in instance.compatibility[lecture_id_1]
            for hall_id_2 in instance.compatibility[lecture_id_2]
        ]
        model.AddAllowedAssignments(
            [hall_assignment[lecture_id_1], hall_assignment[lecture_id_2], z_var],
            tuples,
        )
        objective_terms.append(common_count * z_var)

    if cuts == 3 and z_vars:
        add_cp_extended_strong_distance_propagation(
            model=model,
            instance=instance,
            hall_assignment=hall_assignment,
            z_vars=z_vars,
        )

    model.Minimize(sum(objective_terms) if objective_terms else 0)

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = time_limit
    solver.parameters.num_search_workers = thread_limit
    solver.parameters.random_seed = instance.seed
    solver.parameters.log_search_progress = verbose
    solver.parameters.catch_sigint_signal = False
    status = solver.Solve(model)
    wall_seconds = time.perf_counter() - wall_start

    status_name = status_name_from_cp_sat(status)

    objective_value = None
    lower_bound = safe_float(solver.BestObjectiveBound())
    assignment_by_lecture = None
    if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        objective_value = safe_float(solver.ObjectiveValue())
        assignment_by_lecture = {
            lecture.lecture_id: solver.Value(hall_assignment[lecture.lecture_id])
            for lecture in instance.lectures
        }

    return {
        "solver_family": "OR_TOOLS",
        "formulation": "cp_sat",
        "status": status_name,
        "objective_value": objective_value,
        "lower_bound": lower_bound,
        "wall_clock_seconds": wall_seconds,
        "solver_runtime_seconds": safe_float(solver.WallTime()),
        "mip_gap": None,
        "threads": thread_limit,
        "cuts_mode": cuts,
        "error": None,
        "solution": assignment_details_from_map(instance, assignment_by_lecture),
    }


def preprocessing_infeasible_results(selected_model: str | None, reason: str) -> list[dict[str, Any]]:
    thread_limit = gurobi_thread_limit()
    if selected_model is None:
        model_specs = [
            ("GUROBI", "quadratic_miqp"),
            ("GUROBI", "linearized_milp"),
            ("OR_TOOLS", "cp_sat"),
        ]
    elif selected_model == "MIPQ":
        model_specs = [("GUROBI", "quadratic_miqp")]
    elif selected_model == "MIP":
        model_specs = [("GUROBI", "linearized_milp")]
    elif selected_model == "CP":
        model_specs = [("OR_TOOLS", "cp_sat")]
    else:
        model_specs = [("GUROBI", "linearized_root")]

    return [
        {
            "solver_family": solver_family,
            "formulation": formulation,
            "status": "INFEASIBLE",
            "objective_value": None,
            "lower_bound": None,
            "wall_clock_seconds": 0.0,
            "solver_runtime_seconds": None,
            "mip_gap": None,
            "threads": thread_limit,
            "cuts_mode": None,
            "error": reason,
            "solution": None,
        }
        for solver_family, formulation in model_specs
    ]


def build_summary_rows(
    instance: Instance,
    results: list[dict[str, Any]],
    started_at: dt.datetime,
    finished_at: dt.datetime,
    time_limit: float,
    cuts_mode: int,
    cardinality_enabled: bool,
) -> list[dict[str, Any]]:
    total_lecture_length = sum(lecture.duration for lecture in instance.lectures)
    sizes = [lecture.students for lecture in instance.lectures]
    capacities = [hall.capacity for hall in instance.halls]
    common_values = list(instance.common_students.values())
    peak_active_lectures = max(
        (len(active_lectures) for active_lectures in instance.active_lectures_by_slot.values()),
        default=0,
    )
    peak_slot_density = peak_active_lectures / instance.num_halls if instance.num_halls else 0.0
    assignment_penalty_values = [
        penalty
        for hall_penalties in instance.assignment_penalties.values()
        for penalty in hall_penalties.values()
    ]
    decomposition_connected_components = count_decomposition_connected_components(instance)
    candidate_successors = count_candidate_successor_pairs(instance)

    try:
        script_path = Path(__file__)
        script_name = script_path.name
        script_last_modified = dt.datetime.fromtimestamp(script_path.stat().st_mtime).astimezone().isoformat()
    except Exception:
        script_name = "unknown"
        script_last_modified = "unknown"

    compatibility_entries_removed = (
        instance.compatibility_entries_before - instance.compatibility_entries_after
    )
    compatibility_reduction_ratio = (
        compatibility_entries_removed / instance.compatibility_entries_before
        if instance.compatibility_entries_before
        else 0.0
    )

    valid_lower_bounds = [
        float(result["lower_bound"])
        for result in results
        if result.get("lower_bound") is not None and math.isfinite(float(result["lower_bound"]))
    ]
    best_global_lower_bound = max(valid_lower_bounds) if valid_lower_bounds else None

    rows = []
    for result in results:
        obj = result.get("objective_value")
        lb = result.get("lower_bound")
        gap = result.get("mip_gap")
        solution = result.get("solution")
        walking_objective_value = (
            solution["recomputed_walking_objective"] if solution is not None else None
        )
        matching_penalty_value = (
            solution["recomputed_assignment_penalty"] if solution is not None else None
        )
        if gap is None and obj is not None and lb is not None:
            if obj != 0:
                gap = max(0.0, float(obj - lb) / abs(float(obj)))
            else:
                gap = 0.0 if float(lb) >= 0 else float("inf")

        global_gap = None
        if obj is not None and best_global_lower_bound is not None:
            if obj != 0:
                global_gap = max(0.0, float(obj - best_global_lower_bound) / abs(float(obj)))
            else:
                global_gap = 0.0 if float(best_global_lower_bound) >= 0 else float("inf")

        rows.append(
            {
                "experiment_started_at": started_at.isoformat(),
                "experiment_finished_at": finished_at.isoformat(),
                "script_name": script_name,
                "script_last_modified": script_last_modified,
                "host": socket.gethostname(),
                "platform": platform.platform(),
                "python_version": sys.version.split()[0],
                "seed": instance.seed,
                "instance_name": instance.instance_name,
                "instance_family": instance.instance_family,
                "num_halls": instance.num_halls,
                "days_per_week": instance.days_per_week,
                "slots_per_day": instance.slots_per_day,
                "time_horizon_slots": instance.days_per_week * instance.slots_per_day,
                "density_target": instance.density_target,
                "density_actual": instance.density_actual,
                "raw_slot_minutes": instance.raw_slot_minutes,
                "selected_week_index": instance.selected_week_index,
                "week_selection_mode": instance.week_selection_mode,
                "peak_active_lectures": peak_active_lectures,
                "peak_slot_density": peak_slot_density,
                "free_waste_ratio": FREE_WASTE_RATIO,
                "time_limit_seconds": time_limit,
                "linearized_cuts": cuts_mode,
                "cardinality_enabled": cardinality_enabled,
                "num_lectures": len(instance.lectures),
                "total_lecture_length": total_lecture_length,
                "avg_lecture_length": total_lecture_length / len(instance.lectures),
                "min_lecture_length": min(lecture.duration for lecture in instance.lectures),
                "max_lecture_length": max(lecture.duration for lecture in instance.lectures),
                "min_class_size": min(sizes),
                "avg_class_size": sum(sizes) / len(sizes),
                "max_class_size": max(sizes),
                "min_hall_capacity": min(capacities),
                "avg_hall_capacity": sum(capacities) / len(capacities),
                "max_hall_capacity": max(capacities),
                "candidate_successor_pairs": candidate_successors,
                "successor_pairs_with_common_students": len(instance.common_students),
                "decomposition_connected_components": decomposition_connected_components,
                "compatibility_preprocess_mode": instance.compatibility_preprocess_mode,
                "compatibility_entries_before": instance.compatibility_entries_before,
                "compatibility_entries_after": instance.compatibility_entries_after,
                "compatibility_entries_removed": compatibility_entries_removed,
                "compatibility_reduction_ratio": compatibility_reduction_ratio,
                "compatibility_preprocess_subproblems": instance.compatibility_preprocess_subproblems,
                "compatibility_preprocess_wall_seconds": instance.compatibility_preprocess_wall_seconds,
                "compatibility_preprocess_tightened_lectures": instance.compatibility_preprocess_tightened_lectures,
                "compatibility_preprocess_optimal_subproblems": instance.compatibility_preprocess_optimal_subproblems,
                "compatibility_preprocess_nonoptimal_subproblems": instance.compatibility_preprocess_nonoptimal_subproblems,
                "avg_common_students": (
                    sum(common_values) / len(common_values) if common_values else 0.0
                ),
                "total_common_students_weight": sum(common_values),
                "max_assignment_penalty": max(assignment_penalty_values) if assignment_penalty_values else 0,
                "positive_assignment_penalty_entries": sum(
                    1 for penalty in assignment_penalty_values if penalty > 0
                ),
                "assignment_penalty_type": instance.assignment_penalty_type,
                "successor_max_gap_slots": instance.successor_max_gap_slots,
                "successor_max_gap_minutes": instance.successor_max_gap_minutes,
                "successor_gap_inference_mode": instance.successor_gap_inference_mode,
                "capacity_fix_applied": instance.capacity_fix_applied,
                "capacity_fix_changed_lectures": instance.capacity_fix_changed_lectures,
                "capacity_fix_mode": instance.capacity_fix_mode,
                "fixed_input_time_penalty": instance.fixed_input_time_penalty,
                "fixed_input_time_weight": instance.fixed_input_time_weight,
                "fixed_input_weighted_time_penalty": instance.fixed_input_weighted_time_penalty,
                "fixed_input_time_penalty_allocation": instance.fixed_input_time_penalty_allocation,
                "solver_family": result["solver_family"],
                "formulation": result["formulation"],
                "status": result["status"],
                "objective_value": result["objective_value"],
                "total_student_walking_distance": walking_objective_value,
                "matching_penalty": matching_penalty_value,
                "lower_bound": result["lower_bound"],
                "best_global_lower_bound": best_global_lower_bound,
                "wall_clock_seconds": result["wall_clock_seconds"],
                "solver_runtime_seconds": result["solver_runtime_seconds"],
                "mip_gap": result["mip_gap"],
                "optimality_gap": gap,
                "global opt gap": global_gap,
                "threads": result["threads"],
                "error": result["error"],
            }
        )
    return rows


def excel_cell_value(value: Any) -> Any:
    if value is None:
        return None
    if hasattr(value, "item"):
        try:
            value = value.item()
        except Exception:
            pass
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def sheet_is_empty(worksheet: Any) -> bool:
    return worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet["A1"].value is None


def normalized_row(values: list[Any]) -> list[Any]:
    trimmed = list(values)
    while trimmed and trimmed[-1] is None:
        trimmed.pop()
    return trimmed


def last_header_row(worksheet: Any, first_header_cell: str) -> list[Any] | None:
    if sheet_is_empty(worksheet):
        return None
    for row_idx in range(worksheet.max_row, 0, -1):
        row_values = [
            worksheet.cell(row=row_idx, column=column_idx).value
            for column_idx in range(1, worksheet.max_column + 1)
        ]
        row_values = normalized_row(row_values)
        if row_values and row_values[0] == first_header_cell:
            return row_values
    return None


def count_candidate_successor_pairs(instance: Instance) -> int:
    return sum(
        1
        for lecture in instance.lectures
        for follower in instance.lectures
        if lecture.day == follower.day
        and 0 <= follower.start_slot - lecture.end_slot <= instance.successor_max_gap_slots
    )


def unique_sheet_name(workbook: Workbook, base_name: str) -> str:
    candidate = base_name[:31]
    if candidate not in workbook.sheetnames:
        return candidate
    suffix_index = 1
    while True:
        suffix = f"_{suffix_index}"
        candidate = f"{base_name[:31 - len(suffix)]}{suffix}"
        if candidate not in workbook.sheetnames:
            return candidate
        suffix_index += 1


def write_row(worksheet: Any, row_index: int, values: list[Any]) -> None:
    for column_index, value in enumerate(values, start=1):
        worksheet.cell(row=row_index, column=column_index, value=excel_cell_value(value))


def append_dataframe_to_sheet(workbook: Workbook, sheet_name: str, dataframe: pd.DataFrame) -> None:
    worksheet = workbook.create_sheet(title=unique_sheet_name(workbook, sheet_name))
    row_index = 1
    for row in dataframe_to_rows(dataframe, index=False, header=True):
        write_row(worksheet, row_index, list(row))
        row_index += 1


def append_summary_sheet(workbook: Workbook, summary_df: pd.DataFrame) -> None:
    worksheet = workbook["summary"] if "summary" in workbook.sheetnames else workbook.create_sheet("summary")
    header = list(summary_df.columns)
    prior_header = last_header_row(worksheet, str(header[0])) if header else None
    row_index = 1 if sheet_is_empty(worksheet) else worksheet.max_row + 1
    if prior_header != header:
        write_row(worksheet, row_index, header)
        row_index += 1
    for row in summary_df.itertuples(index=False, name=None):
        write_row(worksheet, row_index, list(row))
        row_index += 1


def instance_to_json_dict(instance: Instance) -> dict[str, Any]:
    peak_active_lectures = max(
        (len(active_lectures) for active_lectures in instance.active_lectures_by_slot.values()),
        default=0,
    )
    peak_slot_density = peak_active_lectures / instance.num_halls if instance.num_halls else 0.0
    if instance.assignment_penalty_type == "itc2019_room_penalty":
        assignment_penalty = {"type": "itc2019_room_penalty"}
    else:
        assignment_penalty = {
            "type": "quadratic_wasted_space",
            "free_waste_ratio": FREE_WASTE_RATIO,
        }
    return {
        "seed": instance.seed,
        "instance_name": instance.instance_name,
        "instance_family": instance.instance_family,
        "num_halls": instance.num_halls,
        "slots_per_day": instance.slots_per_day,
        "days_per_week": instance.days_per_week,
        "density_target": instance.density_target,
        "density_actual": instance.density_actual,
        "raw_slot_minutes": instance.raw_slot_minutes,
        "selected_week_index": instance.selected_week_index,
        "week_selection_mode": instance.week_selection_mode,
        "peak_active_lectures": peak_active_lectures,
        "peak_slot_density": peak_slot_density,
        "successor_rule": {
            "max_gap_slots": instance.successor_max_gap_slots,
            "max_gap_minutes": instance.successor_max_gap_minutes,
            "inference_mode": instance.successor_gap_inference_mode,
        },
        "capacity_fix": {
            "applied": instance.capacity_fix_applied,
            "changed_lectures": instance.capacity_fix_changed_lectures,
            "mode": instance.capacity_fix_mode,
        },
        "assignment_penalty": assignment_penalty,
        "fixed_input_penalties": {
            "time_penalty": instance.fixed_input_time_penalty,
            "time_weight": instance.fixed_input_time_weight,
            "weighted_time_penalty": instance.fixed_input_weighted_time_penalty,
            "time_penalty_allocation": instance.fixed_input_time_penalty_allocation,
        },
        "compatibility_preprocessing": {
            "mode": instance.compatibility_preprocess_mode,
            "entries_before": instance.compatibility_entries_before,
            "entries_after": instance.compatibility_entries_after,
            "entries_removed": (
                instance.compatibility_entries_before - instance.compatibility_entries_after
            ),
            "subproblems_solved": instance.compatibility_preprocess_subproblems,
            "wall_clock_seconds": instance.compatibility_preprocess_wall_seconds,
            "tightened_lectures": instance.compatibility_preprocess_tightened_lectures,
            "optimal_subproblems": instance.compatibility_preprocess_optimal_subproblems,
            "nonoptimal_subproblems": instance.compatibility_preprocess_nonoptimal_subproblems,
        },
        "halls": [
            {
                "hall_id": hall.hall_id,
                "hall_name": hall.name,
                "capacity": hall.capacity,
                "x": hall.x,
                "y": hall.y,
            }
            for hall in instance.halls
        ],
        "lectures": [
            {
                "lecture_id": lecture.lecture_id,
                "lecture_name": lecture.name,
                "subject": lecture.subject,
                "study_year": lecture.study_year,
                "course_type": "compulsory" if lecture.is_compulsory else "elective",
                "day": lecture.day,
                "start_slot_in_day": lecture.start_slot_in_day,
                "duration": lecture.duration,
                "start_slot": lecture.start_slot,
                "end_slot": lecture.end_slot,
                "students": lecture.students,
                "compatible_halls": instance.compatibility[lecture.lecture_id],
                "compatible_hall_penalties": [
                    {
                        "hall_id": hall_id,
                        "penalty": instance.assignment_penalties[lecture.lecture_id][hall_id],
                    }
                    for hall_id in instance.compatibility[lecture.lecture_id]
                ],
            }
            for lecture in instance.lectures
        ],
        "successor_pairs": [
            {
                "from_lecture_id": lecture_id_1,
                "to_lecture_id": lecture_id_2,
                "common_students": common_count,
            }
            for (lecture_id_1, lecture_id_2), common_count in sorted(instance.common_students.items())
        ],
        "distances": instance.distances,
        "active_lectures_by_slot": [
            {"slot": slot, "lecture_ids": lecture_ids}
            for slot, lecture_ids in sorted(instance.active_lectures_by_slot.items())
            if lecture_ids
        ],
    }


def build_json_payload(
    instance: Instance,
    results: list[dict[str, Any]],
    summary_rows: list[dict[str, Any]],
    started_at: dt.datetime,
    finished_at: dt.datetime,
    time_limit: float,
    cuts_mode: int,
    cardinality_enabled: bool,
) -> dict[str, Any]:
    return {
        "experiment": {
            "started_at": started_at.isoformat(),
            "finished_at": finished_at.isoformat(),
            "host": socket.gethostname(),
            "platform": platform.platform(),
            "python_version": sys.version,
            "time_limit_seconds": time_limit,
            "linearized_cuts": cuts_mode,
            "cardinality_enabled": cardinality_enabled,
            "compatibility_preprocess_mode": instance.compatibility_preprocess_mode,
        },
        "instance": instance_to_json_dict(instance),
        "results_summary": summary_rows,
        "solutions": results,
    }


def build_instance_json_payload(
    instance: Instance,
    generated_at: dt.datetime,
) -> dict[str, Any]:
    return {
        "export_type": "instance_only",
        "generation": {
            "generated_at": generated_at.isoformat(),
            "host": socket.gethostname(),
            "platform": platform.platform(),
            "python_version": sys.version.split()[0],
            "command_line": sys.argv,
            "compatibility_preprocess_mode": instance.compatibility_preprocess_mode,
        },
        "instance": instance_to_json_dict(instance),
    }


def write_excel(
    output_path: Path,
    summary_rows: list[dict[str, Any]],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    summary_df = pd.DataFrame(summary_rows)

    if output_path.exists():
        workbook = load_workbook(output_path)
    else:
        workbook = Workbook()
        default_sheet = workbook.active
        if default_sheet and sheet_is_empty(default_sheet):
            workbook.remove(default_sheet)

    append_summary_sheet(workbook, summary_df)
    workbook.save(output_path)


def write_json(output_path: Path, payload: dict[str, Any]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2)


def render_text_table(
    headers: list[str],
    rows: list[list[Any]],
    right_align_columns: set[int] | None = None,
) -> str:
    right_align_columns = right_align_columns or set()
    string_rows = [[str(cell) for cell in row] for row in rows]
    widths = [len(header) for header in headers]
    for row in string_rows:
        for index, cell in enumerate(row):
            widths[index] = max(widths[index], len(cell))

    def format_row(row: list[str]) -> str:
        formatted_cells = []
        for index, cell in enumerate(row):
            if index in right_align_columns:
                formatted_cells.append(cell.rjust(widths[index]))
            else:
                formatted_cells.append(cell.ljust(widths[index]))
        return "  ".join(formatted_cells)

    lines = [format_row(headers), format_row(["-" * width for width in widths])]
    lines.extend(format_row(row) for row in string_rows)
    return "\n".join(lines)


def print_instance_console_view(instance: Instance, json_path: Path) -> None:
    hall_map = {hall.hall_id: hall for hall in instance.halls}
    lecture_map = {lecture.lecture_id: lecture for lecture in instance.lectures}
    total_common_students = sum(instance.common_students.values())
    peak_active_lectures = max(
        (len(active_lectures) for active_lectures in instance.active_lectures_by_slot.values()),
        default=0,
    )
    peak_slot_density = peak_active_lectures / instance.num_halls if instance.num_halls else 0.0
    candidate_successors = count_candidate_successor_pairs(instance)
    min_compatibility = min(len(halls) for halls in instance.compatibility.values()) if instance.compatibility else 0
    max_distance = max(max(row) for row in instance.distances) if instance.distances else 0
    max_assignment_penalty = max(
        (
            penalty
            for hall_penalties in instance.assignment_penalties.values()
            for penalty in hall_penalties.values()
        ),
        default=0,
    )

    print("=" * 88)
    print(
        f"Generated instance input | name={instance.instance_name}, "
        f"family={instance.instance_family}, seed={instance.seed}"
    )
    print("=" * 88)
    print(
        "Overview: "
        f"halls={instance.num_halls}, lectures={len(instance.lectures)}, "
        f"slots/day={instance.slots_per_day}, density={instance.density_actual:.3f} "
        f"(target {instance.density_target:.3f}), peak slot density={peak_slot_density:.3f}"
    )
    if instance.raw_slot_minutes > 0:
        print(
            "ITC timing: "
            f"raw slot={instance.raw_slot_minutes:.3f} minutes, "
            f"selected week={instance.selected_week_index + 1} ({instance.week_selection_mode}), "
            f"successor gap <= {instance.successor_max_gap_slots} slots "
            f"({instance.successor_max_gap_minutes:.3f} minutes, {instance.successor_gap_inference_mode})"
        )
        print(
            "ITC capacity fix: "
            f"applied={instance.capacity_fix_applied}, "
            f"changed lectures={instance.capacity_fix_changed_lectures}, "
            f"mode={instance.capacity_fix_mode}"
        )
    print(
        "Structure: "
        f"successor pairs={len(instance.common_students)}/{candidate_successors}, "
        f"total common students={total_common_students}, "
        f"min compatible halls per lecture={min_compatibility}, max distance={max_distance}, "
        f"max assignment penalty={max_assignment_penalty}, peak active lectures={peak_active_lectures}"
    )
    if instance.assignment_penalty_type == "itc2019_room_penalty":
        print("Penalty: assignment penalties are the ITC 2019 class-room penalties from the instance XML.")
    else:
        print(
            "Penalty: "
            f"wasted space is free up to {FREE_WASTE_RATIO:.0%} of hall capacity; "
            "beyond that, the excess empty seats are penalized quadratically."
        )
    if instance.fixed_input_time_penalty != 0 or instance.fixed_input_time_weight != 0:
        print(
            "Fixed input time penalty: "
            f"raw={instance.fixed_input_time_penalty:.6g}, "
            f"weight={instance.fixed_input_time_weight}, "
            f"weighted={instance.fixed_input_weighted_time_penalty:.6g}, "
            f"allocation={instance.fixed_input_time_penalty_allocation}"
        )
    if instance.compatibility_preprocess_mode != "none":
        removed_entries = instance.compatibility_entries_before - instance.compatibility_entries_after
        print(
            "Preprocess: "
            f"mode={instance.compatibility_preprocess_mode}, "
            f"compatibility entries {instance.compatibility_entries_before} -> "
            f"{instance.compatibility_entries_after} (-{removed_entries}), "
            f"subproblems={instance.compatibility_preprocess_subproblems}, "
            f"tightened lectures={instance.compatibility_preprocess_tightened_lectures}, "
            f"optimal/nonoptimal="
            f"{instance.compatibility_preprocess_optimal_subproblems}/"
            f"{instance.compatibility_preprocess_nonoptimal_subproblems}, "
            f"wall={instance.compatibility_preprocess_wall_seconds:.3f}s"
        )
    if any(len(hall_ids) == 0 for hall_ids in instance.compatibility.values()):
        print("Feasibility: preprocessing detected infeasibility because some lecture has no compatible hall left.")
    print(f"JSON written to: {json_path}")
    print("")

    hall_rows = [
        [
            hall.hall_id,
            hall.name,
            hall.capacity,
            f"{hall.x:.3f}",
            f"{hall.y:.3f}",
        ]
        for hall in instance.halls
    ]
    print("Halls")
    print(
        render_text_table(
            headers=["id", "name", "capacity", "x", "y"],
            rows=hall_rows,
            right_align_columns={0, 2, 3, 4},
        )
    )
    print("")

    lecture_rows = []
    for lecture in instance.lectures:
        slot_end_in_day = lecture.start_slot_in_day + lecture.duration - 1
        compatible_halls = ",".join(
            hall_map[hall_id].name for hall_id in instance.compatibility[lecture.lecture_id]
        )
        lecture_rows.append(
            [
                lecture.lecture_id,
                lecture.name,
                lecture.subject,
                f"Y{lecture.study_year}",
                "compulsory" if lecture.is_compulsory else "elective",
                lecture.day + 1,
                f"{lecture.start_slot_in_day}-{slot_end_in_day}",
                lecture.duration,
                lecture.students,
                compatible_halls,
            ]
        )
    print("Lectures")
    print(
        render_text_table(
            headers=["id", "name", "subject", "year", "type", "day", "slots", "dur", "students", "compatible"],
            rows=lecture_rows,
            right_align_columns={0, 5, 7, 8},
        )
    )
    print("")

    print("Successor pairs")
    if instance.common_students:
        successor_rows = [
            [
                lecture_id_1,
                lecture_map[lecture_id_1].name,
                lecture_id_2,
                lecture_map[lecture_id_2].name,
                common_count,
            ]
            for (lecture_id_1, lecture_id_2), common_count in sorted(instance.common_students.items())
        ]
        print(
            render_text_table(
                headers=["from", "from_name", "to", "to_name", "common_students"],
                rows=successor_rows,
                right_align_columns={0, 2, 4},
            )
        )
    else:
        print("(none)")
    print("")

    distance_rows = []
    hall_headers = ["from/to"] + [hall.name for hall in instance.halls]
    for hall in instance.halls:
        distance_rows.append([hall.name] + instance.distances[hall.hall_id])
    print("Distance matrix")
    print(
        render_text_table(
            headers=hall_headers,
            rows=distance_rows,
            right_align_columns=set(range(1, len(hall_headers))),
        )
    )


def print_console_summary(output_path: Path, summary_rows: list[dict[str, Any]]) -> None:
    print(f"Results written to: {output_path}")
    print("")
    for row in summary_rows:
        gap_str = str(row['optimality_gap']) if row['optimality_gap'] is None else f"{row['optimality_gap']:.2%}"
        global_gap_str = (
            str(row['global opt gap'])
            if row['global opt gap'] is None
            else f"{row['global opt gap']:.2%}"
        )
        print(
            f"{row['solver_family']:>8} | {row['formulation']:<16} | "
            f"status={row['status']:<12} | obj={row['objective_value']} | "
            f"lb={row['lower_bound']} | gap={gap_str} | global={global_gap_str} | "
            f"wall={row['wall_clock_seconds']:.3f}s"
        )


def load_run_instances(args: argparse.Namespace) -> list[tuple[str, Instance]]:
    if args.source == "synthetic":
        try:
            seeds = parse_seed_range(args.seed)
        except ValueError as error:
            raise SystemExit(str(error))
        run_instances: list[tuple[str, Instance]] = []
        for seed in seeds:
            try:
                instance = build_synthetic_instance(
                    num_halls=args.num_halls,
                    slots_per_day=args.slots_per_day,
                    seed=seed,
                    density=args.density,
                )
            except ValueError as error:
                raise SystemExit(str(error))
            run_instances.append((f"seed{seed}", instance))
        return run_instances

    if args.source == "itc2019":
        try:
            instances = load_itc2019_day_instances(
                args.itc_instance,
                solution=args.itc_solution,
                week_index=args.itc_week_index,
                source_day=args.itc_day,
                short_break_slots=args.itc_short_break_slots,
                capacity_fix=args.itc_capacity_fix,
            )
        except ValueError as error:
            raise SystemExit(str(error))
        if not instances:
            raise SystemExit("No active ITC 2019 day instances were loaded.")
        return [(instance.instance_name, instance) for instance in instances]

    try:
        instances = load_lancs_yr23_term_instances(
            args.itc_instance,
            source_day=args.itc_day,
            short_break_slots=args.itc_short_break_slots,
            capacity_fix=args.itc_capacity_fix,
        )
    except ValueError as error:
        raise SystemExit(str(error))
    if not instances:
        raise SystemExit("No active lancs_yr23 day instances were loaded.")
    return [(instance.instance_name, instance) for instance in instances]


def main() -> None:
    args = parse_args()
    run_instances = load_run_instances(args)
    output_path = ensure_output_path(args.output)

    started_at = dt.datetime.now().astimezone()
    base_run_tag = started_at.strftime("%Y%m%d_%H%M%S")
    
    all_summary_rows = []

    for index, (label, instance) in enumerate(run_instances):
        run_tag = f"{base_run_tag}_{label}"
        instance, preprocessing_infeasible = apply_compatibility_preprocessing(
            instance,
            mode=args.compatibility_preprocess,
        )

        if args.instance_only:
            generated_at = dt.datetime.now().astimezone()
            json_path = build_instance_json_path(output_path, run_tag)
            payload = build_instance_json_payload(instance=instance, generated_at=generated_at)
            write_json(json_path, payload)
            if index > 0:
                print("")
            print_instance_console_view(instance, json_path)
            continue

        if preprocessing_infeasible:
            results = preprocessing_infeasible_results(
                args.model,
                reason=(
                    "Compatibility preprocessing proved that no feasible hall assignment "
                    "exists under the hard assignment and overlap constraints."
                ),
            )
        elif args.model is None:
            results = [
                solve_gurobi_quadratic(
                    instance,
                    args.time_limit,
                    verbose=not args.quiet,
                    cardinality=args.cardinality,
                ),
                solve_gurobi_linearized(
                    instance,
                    args.time_limit,
                    cuts=args.cuts,
                    verbose=not args.quiet,
                    cardinality=args.cardinality,
                ),
                solve_cp_sat(
                    instance,
                    args.time_limit,
                    verbose=not args.quiet,
                    cuts=args.cuts,
                    cardinality=args.cardinality,
                ),
            ]
        elif args.model == "MIPQ":
            results = [
                solve_gurobi_quadratic(
                    instance,
                    args.time_limit,
                    verbose=not args.quiet,
                    cardinality=args.cardinality,
                ),
            ]
        elif args.model == "MIP":
            results = [
                solve_gurobi_linearized(
                    instance,
                    args.time_limit,
                    cuts=args.cuts,
                    verbose=not args.quiet,
                    cardinality=args.cardinality,
                ),
            ]
        elif args.model == "CP":
            results = [
                solve_cp_sat(
                    instance,
                    args.time_limit,
                    verbose=not args.quiet,
                    cuts=args.cuts,
                    cardinality=args.cardinality,
                ),
            ]
        else:
            results = [
                solve_gurobi_linearized_root(
                    instance,
                    args.time_limit,
                    cuts=args.cuts,
                    verbose=not args.quiet,
                    cardinality=args.cardinality,
                ),
            ]

        finished_at = dt.datetime.now().astimezone()
        summary_rows = build_summary_rows(
            instance=instance,
            results=results,
            started_at=started_at,
            finished_at=finished_at,
            time_limit=args.time_limit,
            cuts_mode=args.cuts,
            cardinality_enabled=args.cardinality,
        )
        all_summary_rows.extend(summary_rows)
        write_excel(output_path, summary_rows)
        
        if args.save_json:
            json_path = build_json_path(output_path, run_tag)
            payload = build_json_payload(
                instance=instance,
                results=results,
                summary_rows=summary_rows,
                started_at=started_at,
                finished_at=finished_at,
                time_limit=args.time_limit,
                cuts_mode=args.cuts,
                cardinality_enabled=args.cardinality,
            )
            write_json(json_path, payload)
            print(f"JSON written to: {json_path}")

    if not args.instance_only:
        print_console_summary(output_path, all_summary_rows)


if __name__ == "__main__":
    main()
